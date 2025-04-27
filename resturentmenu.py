from tkinter import *
from tkinter import messagebox
from fpdf import FPDF
import datetime
import os
import csv
from openpyxl import Workbook, load_workbook

# Menu data
menu = {
    'pizza': 40,
    'pasta': 50,
    'burger': 60,
    'salad': 70,
    'coffee': 80,
    'full plate veg': 350,
    'breakfast non-veg': 180,
    'full plate non-veg': 500,
    'breakfast veg': 120,
    'cold coffee': 60,
    'chilli patato': 70,
    'chiken chilli': 120,
    'tandur chpati': 20,
    'full plate momo': 70,
    'half plate momo': 40
}

order_total = 0
order_list = []
password = "PYTHON9060"

def add_item():
    item = item_entry.get().lower()
    if item in menu:
        order_list.append(item)
        update_order()
        item_entry.delete(0, END)
    else:
        messagebox.showerror("Item Not Found", f"'{item}' is not available.")

def update_order():
    global order_total
    order_total = sum([menu[i] for i in order_list])
    order_label.config(
        text=f"Items Ordered:\n{', '.join(order_list) if order_list else 'None'}\n{"-"*40}\nTotal: Rs {order_total}\n{"-"*40}"
    )

def save_to_csv(date_str, name, mobile, items, total):
    with open("order_history.csv", "a", newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow([date_str, name, mobile, items, total])

def save_to_excel(date_str, name, mobile, items, total):
    filename = "order_history.xlsx"
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.append(["Date", "Name", "Mobile", "Items", "Total"])
    else:
        wb = load_workbook(filename)
        ws = wb.active

    ws.append([date_str, name, mobile, items, total])
    wb.save(filename)

def finish_order():
    name = name_entry.get()
    mobile = mobile_entry.get()

    if not name or not mobile:
        messagebox.showwarning("Missing Info", "Please enter your name and mobile number.")
        return

    def check_password():
        if pass_entry.get() == password:
            date_str = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            item_list = ', '.join(order_list)

            # --- Create PDF Receipt ---
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", size=12)

            pdf.cell(200, 10, txt="Python Restaurant Receipt", ln=True, align="C")
            pdf.cell(200, 10, txt=f"Date: {date_str}", ln=True, align="C")
            pdf.ln(10)

            pdf.cell(200, 10, txt=f"Name: {name}", ln=True)
            pdf.cell(200, 10, txt=f"Mobile: {mobile}", ln=True)
            pdf.ln(10)

            pdf.cell(200, 10, txt="Items Ordered:", ln=True)
            for item in order_list:
                pdf.cell(200, 10, txt=f"{item.title()} - Rs {menu[item]}", ln=True)

            pdf.ln(5)
            pdf.cell(200, 10, txt=f"\n{"-"*40}\nTotal Amount: Rs {order_total}\n{"-"*40}", ln=True)

            filename = f"{name.replace(' ', '_')}_receipt.pdf"
            pdf.output(filename)

            # Auto-open PDF
            try:
                if os.name == 'nt':
                    os.startfile(filename)
                elif os.name == 'posix':
                    os.system(f'open "{filename}"')
            except Exception as e:
                messagebox.showwarning("Open Failed", f"PDF saved but could not be opened.\n{e}")

            # --- Save to CSV and Excel ---
            save_to_csv(date_str, name, mobile, item_list, order_total)
            save_to_excel(date_str, name, mobile, item_list, order_total)

            messagebox.showinfo("Order Complete", f"Thank you {name}!\nOrder saved to PDF, CSV, and Excel.")

            pass_window.destroy()
            root.quit()
        else:
            messagebox.showerror("Error", "Invalid password. Try again.")
            pass_entry.delete(0, END)
            pass_entry.focus_set()

    pass_window = Toplevel(root)
    pass_window.title("Password Required")
    pass_window.geometry("300x150")
    pass_window.configure(bg="#fff8f0")

    Label(pass_window, text="Enter counter password:", bg="#fff8f0", font="Arial 11").pack(pady=10)
    pass_entry = Entry(pass_window, show="*", font="Arial 11")
    pass_entry.pack()
    pass_entry.focus_set()
    Button(pass_window, text="Submit", command=check_password, bg="#4caf50", fg="white", font="Arial 10 bold").pack(pady=10)

# GUI setup
root = Tk()
root.geometry("700x750")
root.title("🍽️ Python Restaurant")
root.configure(bg="#f0f0f0")

# Header
Label(root, text="Welcome to Python Restaurant", font="Helvetica 20 bold", bg="#f0f0f0", fg="#2e3f50").pack(pady=10)

# Menu display
Label(root, text="Menu", font="Helvetica 16 bold underline", bg="#f0f0f0", fg="#444").pack(pady=10)
menu_text = "\n".join([f"{item.title()}: Rs {price}" for item, price in menu.items()])
menu_frame = Frame(root, bg="#ffffff", bd=2, relief=RIDGE)
menu_frame.pack(pady=5, padx=10)
Label(menu_frame, text=menu_text, justify=LEFT, bg="#ffffff", font="Arial 11").pack(padx=10, pady=10)

# Name and Mobile entry
form_frame = Frame(root, bg="#f0f0f0")
form_frame.pack(pady=5)

Label(form_frame, text="Name:", bg="#f0f0f0", font="Arial 12").grid(row=0, column=0, padx=5, pady=5, sticky=E)
name_entry = Entry(form_frame, font="Arial 12", width=30)
name_entry.grid(row=0, column=1, pady=5)

Label(form_frame, text="Mobile No:", bg="#f0f0f0", font="Arial 12").grid(row=1, column=0, padx=5, pady=5, sticky=E)
mobile_entry = Entry(form_frame, font="Arial 12", width=30)
mobile_entry.grid(row=1, column=1, pady=5)

# Item input
Label(root, text="Enter item name to order:", font="Arial 12", bg="#f0f0f0").pack(pady=(10, 2))
item_entry = Entry(root, width=30, font="Arial 12")
item_entry.pack()

# Buttons
button_frame = Frame(root, bg="#f0f0f0")
button_frame.pack(pady=10)

add_button = Button(button_frame, text="Add to Order", command=add_item, bg="#2196f3", fg="white", font="Arial 11 bold", width=15)
add_button.grid(row=0, column=0, padx=5)

finish_button = Button(button_frame, text="Finish Order", command=finish_order, bg="#4caf50", fg="white", font="Arial 11 bold", width=15)
finish_button.grid(row=0, column=1, padx=5)

# Order summary
order_label = Label(root, text="Items Ordered:\nNone\n\nTotal: Rs 0", font="Arial 12", bg="#f0f0f0", fg="#333", justify=LEFT)
order_label.pack(pady=20)

# Run the app
root.mainloop()
